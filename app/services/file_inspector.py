from __future__ import annotations

import csv
import io
import json
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.file_models import HojaInspeccion, ReporteInspeccion, inferir_metadata_desde_nombre
from app.services.format_detector import FormatoArchivo, detectar_formato, es_ooxml_zip
from app.services.scotiabank_parser import (
    construir_hoja_inspeccion,
    es_estado_cuenta_scotiabank,
    parsear_estado_cuenta_scotiabank,
)


def inspeccionar_archivo(content: bytes, nombre_archivo: str) -> ReporteInspeccion:
    metadata = inferir_metadata_desde_nombre(nombre_archivo)
    formato = detectar_formato(content, nombre_archivo)
    reporte = ReporteInspeccion(
        archivo=nombre_archivo,
        formato_detectado=formato.value,
        hojas_detectadas=[],
        hojas=[],
        banco_inferido=metadata.get("banco_inferido"),
        tipo_fuente_inferido=metadata.get("tipo_fuente_inferido"),
        subtipo_movimiento=metadata.get("subtipo_movimiento"),
        fecha_referencial=metadata.get("fecha_referencial"),
    )

    if formato == FormatoArchivo.PDF:
        reporte.errores_lectura.append("Formato PDF no soportado en V1.")
        return reporte

    if formato == FormatoArchivo.DESCONOCIDO:
        reporte.errores_lectura.append("No se pudo detectar un formato soportado (.csv, .xlsx, .xls).")
        return reporte

    if formato in (FormatoArchivo.XLS_BINARIO, FormatoArchivo.XLSX_OOXML) and es_estado_cuenta_scotiabank(
        content, nombre_archivo
    ):
        return _inspeccionar_scotiabank(content, nombre_archivo, metadata)

    try:
        if formato == FormatoArchivo.CSV:
            hoja = _inspeccionar_csv(content)
            reporte.hojas = [hoja]
            reporte.hojas_detectadas = [hoja.nombre]
        elif formato == FormatoArchivo.XLSX_OOXML:
            if not es_ooxml_zip(content):
                reporte.errores_lectura.append("Archivo ZIP detectado pero no parece Excel OOXML válido.")
                return reporte
            hojas = _inspeccionar_ooxml(content)
            reporte.hojas = hojas
            reporte.hojas_detectadas = [h.nombre for h in hojas]
        elif formato == FormatoArchivo.XLS_BINARIO:
            hojas = _inspeccionar_xls_binario(content)
            reporte.hojas = hojas
            reporte.hojas_detectadas = [h.nombre for h in hojas]
    except Exception as exc:
        reporte.errores_lectura.append(f"Error al leer archivo: {exc}")

    return reporte


def _inspeccionar_scotiabank(content: bytes, nombre_archivo: str, metadata: dict) -> ReporteInspeccion:
    meta_parseo, movimientos = parsear_estado_cuenta_scotiabank(content)
    hoja_dict = construir_hoja_inspeccion(movimientos)
    hoja = HojaInspeccion(**hoja_dict)

    return ReporteInspeccion(
        archivo=nombre_archivo,
        formato_detectado=FormatoArchivo.XLS_BINARIO.value,
        hojas_detectadas=[hoja.nombre],
        hojas=[hoja],
        banco_inferido=meta_parseo.get("banco_inferido") or metadata.get("banco_inferido"),
        tipo_fuente_inferido=meta_parseo.get("tipo_fuente_inferido") or metadata.get("tipo_fuente_inferido"),
        subtipo_movimiento=meta_parseo.get("subtipo_movimiento") or metadata.get("subtipo_movimiento"),
        fecha_referencial=meta_parseo.get("fecha_referencial") or metadata.get("fecha_referencial"),
        errores_lectura=[] if movimientos else ["No se encontraron movimientos en el estado de cuenta Scotiabank."],
    )


def _detectar_fila_encabezado(filas: list[list[Any]], max_busqueda: int = 30) -> tuple[int | None, list[str]]:
    mejor_fila = None
    mejor_puntaje = 0
    mejor_encabezados: list[str] = []

    limite = min(len(filas), max_busqueda)
    for idx in range(limite):
        fila = filas[idx]
        textos = [_celda_a_texto(c) for c in fila]
        no_vacios = [t for t in textos if t]
        if len(no_vacios) < 2:
            continue
        puntaje = len(no_vacios)
        if puntaje > mejor_puntaje:
            mejor_puntaje = puntaje
            mejor_fila = idx + 1
            mejor_encabezados = [_normalizar_encabezado(t) for t in textos]

    return mejor_fila, mejor_encabezados


def _construir_hoja(nombre: str, filas: list[list[Any]]) -> HojaInspeccion:
    filas_no_vacias = [f for f in filas if any(_celda_a_texto(c) for c in f)]
    fila_encabezado, columnas = _detectar_fila_encabezado(filas_no_vacias)
    max_cols = max((len(f) for f in filas_no_vacias), default=0)

    filas_ejemplo: list[dict] = []
    if fila_encabezado and columnas:
        inicio = fila_encabezado
        for fila in filas_no_vacias[inicio : inicio + 5]:
            registro = {}
            for idx, col in enumerate(columnas):
                if not col:
                    col = f"col_{idx + 1}"
                valor = fila[idx] if idx < len(fila) else ""
                registro[col] = _celda_a_texto(valor)
            if any(registro.values()):
                filas_ejemplo.append(registro)

    ultima_fila = len(filas_no_vacias)
    rango = f"A1:{get_column_letter(max(max_cols, 1))}{max(ultima_fila, 1)}"

    return HojaInspeccion(
        nombre=nombre,
        filas_totales=len(filas_no_vacias),
        columnas_totales=max_cols,
        rango_usado=rango,
        fila_encabezado_probable=fila_encabezado,
        columnas_detectadas=[c for c in columnas if c],
        filas_ejemplo=filas_ejemplo,
    )


def _inspeccionar_csv(content: bytes) -> HojaInspeccion:
    df, _ = _leer_csv(content)
    filas = df.fillna("").values.tolist()
    return _construir_hoja("CSV", filas)


def _inspeccionar_ooxml(content: bytes) -> list[HojaInspeccion]:
    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    hojas: list[HojaInspeccion] = []
    for nombre in wb.sheetnames:
        ws = wb[nombre]
        filas = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
        if any(any(_celda_a_texto(c) for c in fila) for fila in filas):
            hojas.append(_construir_hoja(nombre, filas))
    wb.close()
    return hojas


def _inspeccionar_xls_binario(content: bytes) -> list[HojaInspeccion]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("Lector xlrd no disponible para .xls binario.") from exc

    libro = xlrd.open_workbook(file_contents=content)
    hojas: list[HojaInspeccion] = []
    for hoja in libro.sheets():
        filas = [hoja.row_values(i) for i in range(hoja.nrows)]
        if any(any(_celda_a_texto(c) for c in fila) for fila in filas):
            hojas.append(_construir_hoja(hoja.name, filas))
    return hojas


def extraer_filas_raw(content: bytes, nombre_archivo: str, reporte: ReporteInspeccion) -> list[dict]:
    """Extrae filas tabulares como dict usando encabezados detectados por hoja."""
    if es_estado_cuenta_scotiabank(content, nombre_archivo):
        _, movimientos = parsear_estado_cuenta_scotiabank(content)
        resultado: list[dict] = []
        for mov in movimientos:
            registro = {k: v for k, v in mov.items() if not k.startswith("_")}
            registro["_hoja_origen"] = mov.get("_hoja_origen", "estado_cta_trj")
            registro["_fila_origen"] = int(mov.get("_fila_origen", 0))
            resultado.append(registro)
        return resultado

    formato = FormatoArchivo(reporte.formato_detectado)
    resultado: list[dict] = []

    if formato == FormatoArchivo.CSV:
        df, _ = _leer_csv(content)
        hoja = reporte.hojas[0] if reporte.hojas else None
        if not hoja or not hoja.fila_encabezado_probable:
            return resultado
        filas = df.fillna("").values.tolist()
        encabezados = _encabezados_desde_hoja(hoja, filas[hoja.fila_encabezado_probable - 1])
        for offset, fila in enumerate(filas[hoja.fila_encabezado_probable :], start=hoja.fila_encabezado_probable + 1):
            registro = _fila_a_dict(encabezados, fila)
            if _fila_tiene_datos(registro):
                registro["_hoja_origen"] = "CSV"
                registro["_fila_origen"] = offset
                resultado.append(registro)
        return resultado

    if formato == FormatoArchivo.XLSX_OOXML:
        wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
        for hoja_info in reporte.hojas:
            if not hoja_info.fila_encabezado_probable:
                continue
            ws = wb[hoja_info.nombre]
            filas = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
            encabezados = _encabezados_desde_hoja(hoja_info, filas[hoja_info.fila_encabezado_probable - 1])
            for offset, fila in enumerate(
                filas[hoja_info.fila_encabezado_probable :],
                start=hoja_info.fila_encabezado_probable + 1,
            ):
                registro = _fila_a_dict(encabezados, fila)
                if _fila_tiene_datos(registro):
                    registro["_hoja_origen"] = hoja_info.nombre
                    registro["_fila_origen"] = offset
                    resultado.append(registro)
        wb.close()
        return resultado

    if formato == FormatoArchivo.XLS_BINARIO:
        import xlrd

        libro = xlrd.open_workbook(file_contents=content)
        hoja_map = {h.nombre: h for h in reporte.hojas}
        for sheet in libro.sheets():
            hoja_info = next((h for h in reporte.hojas if h.nombre == sheet.name), None)
            if not hoja_info or not hoja_info.fila_encabezado_probable:
                continue
            filas = [sheet.row_values(i) for i in range(sheet.nrows)]
            encabezados = _encabezados_desde_hoja(hoja_info, filas[hoja_info.fila_encabezado_probable - 1])
            for offset, fila in enumerate(
                filas[hoja_info.fila_encabezado_probable :],
                start=hoja_info.fila_encabezado_probable + 1,
            ):
                registro = _fila_a_dict(encabezados, fila)
                if _fila_tiene_datos(registro):
                    registro["_hoja_origen"] = sheet.name
                    registro["_fila_origen"] = offset
                    resultado.append(registro)
        return resultado

    return resultado


def _leer_csv(content: bytes) -> tuple[pd.DataFrame, str]:
    ultimo_error: Exception | None = None
    for encoding in ("utf-8-sig", "latin-1", "cp1252"):
        for sep in (None, ",", ";", "\t", "|"):
            try:
                buffer = io.BytesIO(content)
                df = pd.read_csv(buffer, sep=sep, engine="python", encoding=encoding, dtype=str)
                if df.shape[1] >= 2:
                    return df.fillna(""), encoding
            except Exception as exc:
                ultimo_error = exc
    raise ValueError(f"No se pudo leer CSV: {ultimo_error}")


def _encabezados_desde_hoja(hoja: HojaInspeccion, fila_encabezado: list[Any]) -> list[str]:
    if hoja.columnas_detectadas:
        return hoja.columnas_detectadas
    return [_normalizar_encabezado(_celda_a_texto(c)) or f"col_{idx + 1}" for idx, c in enumerate(fila_encabezado)]


def _fila_a_dict(encabezados: list[str], fila: list[Any]) -> dict:
    registro: dict[str, str] = {}
    for idx, col in enumerate(encabezados):
        clave = col or f"col_{idx + 1}"
        registro[clave] = _celda_a_texto(fila[idx] if idx < len(fila) else "")
    return registro


def _fila_tiene_datos(registro: dict) -> bool:
    return any(v for k, v in registro.items() if not k.startswith("_") and v)


def _celda_a_texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _normalizar_encabezado(texto: str) -> str:
    return " ".join(texto.split())


def reporte_a_json(reporte: ReporteInspeccion) -> str:
    return json.dumps(reporte.to_dict(), ensure_ascii=False, indent=2)
